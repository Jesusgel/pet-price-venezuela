"""Script de carga única: lee productos.xlsx y los inserta en la base de datos.

Columnas usadas:
  - 'Nombre Producto' -> name
  - 'Unidad'          -> unit
  - 'Precio Venta ($)'-> price_usd

Columnas ignoradas: 'Cantidad', 'Precio Compra ($)'
Campos sin datos en el Excel: category='', brand=None, weight_kg=None
"""

import asyncio
import os
import sys
from decimal import Decimal, InvalidOperation
from pathlib import Path

import openpyxl

# Permite importar `app` desde cualquier contexto
sys.path.append(str(Path(__file__).resolve().parent.parent))

from app.core.database import async_session_maker  # noqa: E402
from app.models.product import Product  # noqa: E402
from sqlmodel import select  # noqa: E402

# Ruta al archivo Excel en la raíz del proyecto
XLSX_PATH = Path(__file__).resolve().parent.parent.parent / "productos.xlsx"

# Columnas esperadas en el Excel
COL_NAME = "Nombre Producto"
COL_UNIT = "Unidad"
COL_PRICE_SELL = "Precio Venta ($)"


def parse_rows(ws) -> list[dict]:
    """Parsea las filas del Excel y retorna una lista de dicts listos para Product."""
    headers = [cell.value for cell in ws[1]]

    try:
        idx_name = headers.index(COL_NAME)
        idx_unit = headers.index(COL_UNIT)
        idx_price = headers.index(COL_PRICE_SELL)
    except ValueError as e:
        raise ValueError(f"Columna no encontrada en el Excel: {e}") from e

    products: list[dict] = []
    for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        name = row[idx_name]
        unit = row[idx_unit]
        price_raw = row[idx_price]

        # Saltar filas vacías
        if not name:
            continue

        try:
            price_usd = Decimal(str(price_raw))
        except (InvalidOperation, TypeError):
            print(f"[WARN] Fila {row_num}: precio invalido '{price_raw}', se omite.")
            continue

        products.append(
            {
                "name": str(name).strip(),
                "unit": str(unit).strip() if unit else "unidad",
                "price_usd": price_usd,
            }
        )

    return products


async def seed_from_xlsx() -> None:
    """Lee el Excel e inserta los productos en la base de datos."""
    if not XLSX_PATH.exists():
        print(f"[ERROR] Archivo no encontrado: {XLSX_PATH}")
        sys.exit(1)

    wb = openpyxl.load_workbook(XLSX_PATH)
    ws = wb.active
    rows = parse_rows(ws)

    if not rows:
        print("[WARN] No se encontraron filas validas en el Excel.")
        return

    async with async_session_maker() as session:
        # Verificar duplicados por nombre
        statement = select(Product.name)
        result = await session.exec(statement)
        existing_names: set[str] = {r for r in result.all()}

        new_products: list[Product] = []
        skipped = 0

        for data in rows:
            if data["name"] in existing_names:
                print(f"   [SKIP] Ya existe: {data['name']}")
                skipped += 1
                continue

            new_products.append(
                Product(
                    name=data["name"],
                    unit=data["unit"],
                    price_usd=data["price_usd"],
                    category="",   # Sin categoría por ahora
                    brand=None,    # Sin marca por ahora
                    weight_kg=None,
                    is_active=True,
                )
            )

        if not new_products:
            print("[INFO] Todos los productos ya existen en la base de datos.")
            return

        session.add_all(new_products)
        await session.commit()

        print(f"\n[OK] {len(new_products)} productos insertados correctamente.")
        if skipped:
            print(f"   {skipped} productos omitidos (duplicados).")


if __name__ == "__main__":
    asyncio.run(seed_from_xlsx())
